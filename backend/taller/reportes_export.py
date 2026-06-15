"""
CU18: Plantillas de exportación (CSV, Excel, PDF) para Generar Reportes.
Incluye encabezado con el logo y nombre del taller, título del reporte,
filtros aplicados y una tabla con columnas/formatos legibles en español.
"""
import csv
import io
from pathlib import Path

from django.http import HttpResponse

NOMBRE_EMPRESA = 'Taller de Motos La Roca'
LOGO_PATH = Path(__file__).resolve().parent / 'static' / 'branding' / 'logo_la_roca.png'
LOGO_RATIO = 1640 / 152  # ancho / alto del archivo logo_la_roca.png

# Etiqueta y tipo de formato ('text' | 'int' | 'money' | 'percent') por columna
COLUMNAS_INFO = {
    'periodo': ('Periodo', 'text'),
    'ordenes': ('# Órdenes', 'int'),
    'ingreso_bruto': ('Ingreso Bruto', 'money'),
    'impuesto': ('Impuesto', 'money'),
    'ingreso_neto': ('Ingreso Neto', 'money'),
    'servicio': ('Servicio', 'text'),
    'veces_realizado': ('Veces Realizado', 'int'),
    'ingreso_total': ('Ingreso Total', 'money'),
    'porcentaje': ('% del Total', 'percent'),
    'repuesto': ('Repuesto', 'text'),
    'cantidad_vendida': ('Cantidad Vendida', 'int'),
    'stock_actual': ('Stock Actual', 'int'),
    'cliente': ('Cliente', 'text'),
    'cedula': ('Cédula', 'text'),
    'cantidad_servicios': ('# Servicios', 'int'),
    'total_gastado': ('Total Gastado', 'money'),
    'estado': ('Estado', 'text'),
    'cantidad': ('Cantidad', 'int'),
    'producto': ('Producto', 'text'),
    'stock_minimo': ('Stock Mínimo', 'int'),
    'diferencia': ('Diferencia', 'int'),
    'ubicacion': ('Ubicación', 'text'),
}

TITULOS_REPORTE = {
    'ingresos_por_periodo': 'Ingresos por Período',
    'servicios_mas_realizados': 'Servicios Más Realizados',
    'repuestos_mas_vendidos': 'Repuestos Más Vendidos',
    'clientes_frecuentes': 'Clientes Frecuentes',
    'ordenes_por_estado': 'Órdenes por Estado',
    'inventario_critico': 'Inventario Crítico',
}

# Columnas por defecto cuando no hay resultados (para que el encabezado
# de la tabla no quede vacío, p.ej. Inventario crítico sin productos bajo mínimo)
COLUMNAS_POR_TIPO = {
    'ingresos_por_periodo': ['periodo', 'ordenes', 'ingreso_bruto', 'impuesto', 'ingreso_neto'],
    'servicios_mas_realizados': ['servicio', 'veces_realizado', 'ingreso_total', 'porcentaje'],
    'repuestos_mas_vendidos': ['repuesto', 'cantidad_vendida', 'ingreso_total', 'stock_actual'],
    'clientes_frecuentes': ['cliente', 'cedula', 'cantidad_servicios', 'total_gastado'],
    'ordenes_por_estado': ['estado', 'cantidad', 'porcentaje'],
    'inventario_critico': ['producto', 'stock_actual', 'stock_minimo', 'diferencia', 'ubicacion'],
}


def columnas_para(tipo, resultados):
    """Devuelve la lista de columnas a exportar para un tipo de reporte."""
    if resultados:
        return list(resultados[0].keys())
    return COLUMNAS_POR_TIPO.get(tipo, ['categoria', 'descripcion', 'cantidad', 'monto', 'estado'])


def etiqueta_columna(col):
    return COLUMNAS_INFO.get(col, (col.replace('_', ' ').title(), 'text'))[0]


def formato_columna(col):
    return COLUMNAS_INFO.get(col, (col, 'text'))[1]


def formatear_valor(col, val):
    if val is None or val == '':
        return '-'
    tipo_formato = formato_columna(col)
    if tipo_formato == 'money':
        try:
            return f"Bs {float(val):,.2f}"
        except (TypeError, ValueError):
            return str(val)
    if tipo_formato == 'percent':
        try:
            return f"{float(val):.2f}%"
        except (TypeError, ValueError):
            return str(val)
    return str(val)


def construir_subtitulo(filtros):
    partes = [TITULOS_REPORTE.get(filtros.get('tipo'), filtros.get('tipo') or 'Reporte')]
    rango = []
    if filtros.get('fecha_inicio'):
        rango.append(f"Desde {filtros['fecha_inicio']}")
    if filtros.get('fecha_fin'):
        rango.append(f"Hasta {filtros['fecha_fin']}")
    if rango:
        partes.append(' '.join(rango))
    if filtros.get('agrupacion'):
        partes.append(f"Agrupación: {filtros['agrupacion'].capitalize()}")
    partes.append(f"Top: {filtros.get('top', 10)}")
    return ' | '.join(partes)


def generar_csv_reporte(headers, resultados, filtros):
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="reporte.csv"'
    response.write('﻿')  # BOM para que Excel detecte UTF-8 correctamente

    writer = csv.writer(response)
    writer.writerow([NOMBRE_EMPRESA])
    writer.writerow([construir_subtitulo(filtros)])
    writer.writerow([f"Generado: {filtros.get('generado_en', '')}"])
    writer.writerow([])
    writer.writerow([etiqueta_columna(col) for col in headers])
    for row in resultados:
        writer.writerow([row.get(col, '') for col in headers])
    if not resultados:
        writer.writerow(['Sin resultados para los filtros seleccionados.'])
    return response


def generar_excel_reporte(headers, resultados, filtros):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = 'Reporte'
    n_cols = max(len(headers), 1)

    if LOGO_PATH.exists():
        from openpyxl.drawing.image import Image as XLImage
        img = XLImage(str(LOGO_PATH))
        img.width = 220
        img.height = round(220 / LOGO_RATIO)
        ws.add_image(img, 'A1')
        ws.row_dimensions[1].height = round(img.height * 0.75)

    titulo_cell = ws.cell(row=3, column=1, value=NOMBRE_EMPRESA)
    titulo_cell.font = Font(size=14, bold=True, color='C0392B')
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=n_cols)

    subtitulo_cell = ws.cell(row=4, column=1, value=construir_subtitulo(filtros))
    subtitulo_cell.font = Font(size=10, bold=True)
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=n_cols)

    info_cell = ws.cell(row=5, column=1, value=f"Generado: {filtros.get('generado_en', '')}")
    info_cell.font = Font(size=9, italic=True, color='808080')
    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=n_cols)

    header_row = 7
    header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    zebra_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    total_fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for c, col in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=c, value=etiqueta_columna(col))
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    for r, row in enumerate(resultados, start=header_row + 1):
        es_total = str(row.get(headers[0], '')).strip().upper() == 'TOTAL'
        for c, col in enumerate(headers, start=1):
            val = row.get(col)
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
            fmt = formato_columna(col)
            if fmt == 'money' and isinstance(val, (int, float)):
                cell.number_format = '"Bs" #,##0.00'
            elif fmt == 'percent' and isinstance(val, (int, float)):
                cell.number_format = '0.00"%"'
            if es_total:
                cell.font = Font(bold=True)
                cell.fill = total_fill
            elif (r - header_row) % 2 == 0:
                cell.fill = zebra_fill

    if not resultados:
        cell = ws.cell(row=header_row + 1, column=1, value='Sin resultados para los filtros seleccionados.')
        cell.font = Font(italic=True, color='808080')
        ws.merge_cells(start_row=header_row + 1, start_column=1, end_row=header_row + 1, end_column=n_cols)

    for c, col in enumerate(headers, start=1):
        valores = [str(formatear_valor(col, row.get(col))) for row in resultados]
        max_len = max([len(etiqueta_columna(col))] + [len(v) for v in valores] + [8])
        ws.column_dimensions[get_column_letter(c)].width = min(max_len + 4, 40)

    output = io.BytesIO()
    wb.save(output)
    contenido = output.getvalue()

    response = HttpResponse(
        contenido,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="reporte.xlsx"'
    return response, contenido


def generar_pdf_reporte(headers, resultados, filtros):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import landscape, letter
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import Image as RLImage
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    output = io.BytesIO()
    n_cols = max(len(headers), 1)
    pagesize = landscape(letter) if n_cols > 4 else letter
    doc = SimpleDocTemplate(
        output, pagesize=pagesize,
        topMargin=16 * mm, bottomMargin=14 * mm, leftMargin=14 * mm, rightMargin=14 * mm,
    )

    styles = getSampleStyleSheet()
    estilo_titulo = ParagraphStyle('TituloEmpresa', parent=styles['Title'], fontSize=16, textColor=colors.HexColor('#C0392B'), spaceAfter=2)
    estilo_subtitulo = ParagraphStyle('Subtitulo', parent=styles['Heading2'], fontSize=12, spaceAfter=4)
    estilo_info = ParagraphStyle('Info', parent=styles['Normal'], fontSize=8, textColor=colors.grey)

    elementos = []
    if LOGO_PATH.exists():
        ancho = 55 * mm
        elementos.append(RLImage(str(LOGO_PATH), width=ancho, height=ancho / LOGO_RATIO))
        elementos.append(Spacer(1, 6))
    elementos.append(Paragraph(NOMBRE_EMPRESA, estilo_titulo))
    elementos.append(Paragraph(TITULOS_REPORTE.get(filtros.get('tipo'), filtros.get('tipo') or 'Reporte'), estilo_subtitulo))
    elementos.append(Paragraph(construir_subtitulo(filtros), estilo_info))
    elementos.append(Paragraph(f"Generado: {filtros.get('generado_en', '')}", estilo_info))
    elementos.append(Spacer(1, 10))

    tabla_data = [[etiqueta_columna(col) for col in headers]]
    for row in resultados:
        tabla_data.append([formatear_valor(col, row.get(col)) for col in headers])
    if not resultados:
        fila_vacia = ['Sin resultados para los filtros seleccionados.'] + [''] * (n_cols - 1)
        tabla_data.append(fila_vacia)

    tabla = Table(tabla_data, repeatRows=1, hAlign='LEFT')
    estilo_tabla = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2C3E50')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')]),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ])
    for idx, row in enumerate(resultados, start=1):
        if str(row.get(headers[0], '')).strip().upper() == 'TOTAL':
            estilo_tabla.add('FONTNAME', (0, idx), (-1, idx), 'Helvetica-Bold')
            estilo_tabla.add('BACKGROUND', (0, idx), (-1, idx), colors.HexColor('#E8E8E8'))
    tabla.setStyle(estilo_tabla)
    elementos.append(tabla)

    doc.build(elementos)
    contenido = output.getvalue()

    response = HttpResponse(contenido, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte.pdf"'
    return response, contenido
